"""
アンケート解析特化型 GLMプラットフォーム (ロジスティック回帰)
主な改善点:
  - 二項分布（ロジスティック回帰）への固定
  - オッズ比（Odds Ratio）の自動算出と表示
  - 欠損値（無回答）の自動除外と警告表示
  - アンケート向けの可視化（カテゴリ：棒グラフ割合、数値：ロジスティック曲線）
  - 官能評価・消費者アンケート向けサンプルデータへの刷新
"""

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import seaborn as sns
import statsmodels.api as sm
import statsmodels.formula.api as smf
from statsmodels.tools.sm_exceptions import PerfectSeparationError
from matplotlib import font_manager
import io
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage

warnings.filterwarnings('ignore')

# ==========================================
# 定数・設定
# ==========================================
st.set_page_config(
    page_title="アンケート解析 GLMプラットフォーム",
    page_icon="📋",
    layout="wide"
)

ALPHA = 0.05  # 有意水準

# ==========================================
# 日本語フォント設定
# ==========================================
@st.cache_resource
def setup_japanese_font():
    try:
        import japanize_matplotlib
        japanize_matplotlib.japanize()
        return "japanize_matplotlib"
    except ImportError:
        pass

    candidates = [
        'IPAexGothic', 'IPAPGothic', 'Noto Sans CJK JP', 'Noto Sans JP',
        'Hiragino Sans', 'Hiragino Maru Gothic Pro',
        'MS Gothic', 'Yu Gothic', 'Meiryo', 'VL Gothic'
    ]
    available = {f.name for f in font_manager.fontManager.ttflist}
    for font in candidates:
        if font in available:
            plt.rcParams['font.family'] = font
            plt.rcParams['axes.unicode_minus'] = False
            return font

    plt.rcParams['axes.unicode_minus'] = False
    return "fallback"

FONT_STATUS = setup_japanese_font()

# ==========================================
# ユーティリティ関数
# ==========================================
def reset_session():
    keys_to_clear = [
        'analyzed', 'report_images', 'model_result',
        'report_glm', 'report_wald', 'df_eval',
        'eval_col', 'formula', 'target_col', 'factor_cols', 'factor_types'
    ]
    for k in keys_to_clear:
        if k in st.session_state:
            del st.session_state[k]
    st.rerun()

def detect_perfect_separation(df: pd.DataFrame, target_col: str, factor_col: str) -> list:
    problems = []
    unique_vals = df[target_col].dropna().unique()
    is_binary = set(unique_vals).issubset({0, 1, 0.0, 1.0})
    if is_binary:
        for cat, grp in df.groupby(factor_col):
            vals = grp[target_col].dropna()
            if len(vals) > 0 and vals.nunique() == 1:
                problems.append(str(cat))
    return problems

def safe_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors='coerce')

def make_fig_for_category(df, factor, eval_col, target_col):
    """アンケート向け：水準ごとの「1」の割合を棒グラフで表示"""
    prop = df.groupby(factor)[eval_col].mean().sort_values(ascending=False)
    counts = df.groupby(factor)[eval_col].count()
    
    fig_w = max(5, 3 + len(prop) * 0.8)
    fig, ax = plt.subplots(figsize=(fig_w, 4))
    
    sns.barplot(x=prop.index, y=prop.values, ax=ax, color='#4F81BD', alpha=0.8)
    
    for i, p in enumerate(prop.values):
        n_count = counts[prop.index[i]]
        ax.text(i, p + 0.02, f"{p*100:.1f}%\n(n={n_count})", ha='center', va='bottom', fontsize=9)
        
    ax.set_ylim(0, max(1.0, prop.max() * 1.2)) # 上部に余裕を持たせる
    ax.set_xlabel(factor, fontsize=10)
    ax.set_ylabel(f"{target_col} の該当割合", fontsize=10)
    ax.tick_params(axis='x', rotation=30 if len(prop) > 5 else 0)
    plt.tight_layout()
    return fig

def make_fig_for_numeric(df, factor, eval_col, target_col):
    """アンケート向け：ロジスティック曲線の描画を強制"""
    fig, ax = plt.subplots(figsize=(5, 4))
    try:
        sns.regplot(
            x=factor, y=eval_col, data=df, ax=ax,
            scatter_kws={'alpha': 0.3, 'color': '#333333', 's': 30, 'y_jitter': 0.05},
            line_kws={'color': 'crimson', 'linewidth': 2},
            logistic=True # ロジスティック曲線を強制
        )
    except Exception:
        # エラー時は通常の散布図
        sns.scatterplot(x=factor, y=eval_col, data=df, ax=ax, alpha=0.5)
        
    ax.set_xlabel(factor, fontsize=10)
    ax.set_ylabel(f"{target_col} (確率)", fontsize=10)
    ax.set_yticks([0, 0.5, 1])
    plt.tight_layout()
    return fig

def fig_to_bytesio(fig) -> io.BytesIO:
    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', dpi=120)
    buf.seek(0)
    plt.close(fig)
    return buf

def generate_ai_prompt(target_col, formula, report_glm, report_wald):
    try:
        glm_md = report_glm.reset_index().to_markdown(index=False, floatfmt='.4f')
    except Exception:
        glm_md = report_glm.reset_index().to_csv(sep='\t', index=False)

    if report_wald is not None:
        try:
            wald_md = report_wald.reset_index().to_markdown(index=False, floatfmt='.4f')
        except Exception:
            wald_md = report_wald.reset_index().to_csv(sep='\t', index=False)
    else:
        wald_md = "計算不可（データなし）"

    clean_formula = formula.replace('Q(', '').replace(')', '').replace('"', '')

    prompt = f"""あなたは統計解析とマーケティング（消費者アンケート）の専門家です。以下のアンケート調査データにおけるロジスティック回帰（二項分布のGLM）の解析結果を解釈し、実践的なインサイトを提供してください。

## 解析の前提条件
- **目的変数（1=該当, 0=非該当）**: {target_col}
- **モデル式**: `{clean_formula}`
- **有意水準**: α = {ALPHA}

## Wald検定表（質問項目全体の有意性）
{wald_md}

## 回帰係数とオッズ比（水準別の影響度）
{glm_md}

## 解釈の指示
1. **Wald検定表**から、目的変数に対して有意な影響を与えている質問項目や要因を抽出してください。
2. **オッズ比 (Odds Ratio)** に着目し、「ある水準を満たすと、そうでない場合と比べて何倍『{target_col}』になりやすいか」を具体的に解説してください。（オッズ比 > 1 は促進要因、< 1 は阻害要因）
3. **Std.Error やオッズ比が異常に大きい** 項目があれば、回答の偏り（完全分離）や質問間の似すぎ（多重共線性）を疑い、注意を促してください。
4. この結果を踏まえ、商品の改善や次の調査設計に向けたアクションプランを提案してください。
"""
    return prompt

def generate_excel_report(target_col, formula, report_glm, report_wald, report_images: dict) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "アンケート解析レポート"

    title_font   = Font(size=14, bold=True)
    head_font    = Font(bold=True, color="FFFFFF")
    head_fill    = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subhead_font = Font(bold=True, size=11)
    thin_bd = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    sig_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def write_header_row(ws, row, headers, col_start=1):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=col_start + i, value=h)
            c.font = head_font; c.fill = head_fill; c.border = thin_bd
            c.alignment = Alignment(horizontal='center')

    def write_data_cell(ws, row, col, value, highlight=False):
        c = ws.cell(row=row, column=col, value=value)
        c.border = thin_bd
        if highlight: c.fill = sig_fill
        return c

    cur = 1
    ws.cell(row=cur, column=1, value="📋 アンケート解析レポート (ロジスティック回帰)").font = title_font
    cur += 2
    ws.cell(row=cur, column=1, value="【モデル設定】").font = subhead_font
    cur += 1
    ws.cell(row=cur, column=1, value=f"目的変数: {target_col}")
    cur += 1
    clean_f = formula.replace('Q(', '').replace(')', '').replace('"', '')
    ws.cell(row=cur, column=1, value=f"モデル式: {clean_f}")
    cur += 2

    ws.cell(row=cur, column=1, value="【1. Wald検定表 (要因の有意性)】").font = subhead_font
    cur += 1
    if report_wald is not None:
        headers_w = ["要因", "Df", "Chi-Square", "Pr(>Chisq)", "有意"]
        write_header_row(ws, cur, headers_w)
        cur += 1
        for idx, row_data in report_wald.iterrows():
            is_sig = float(row_data['Pr(>Chisq)']) < ALPHA
            write_data_cell(ws, cur, 1, str(idx), highlight=is_sig)
            write_data_cell(ws, cur, 2, int(row_data['Df']) if pd.notna(row_data['Df']) else '', highlight=is_sig)
            write_data_cell(ws, cur, 3, round(float(row_data['Chi-Square']), 3) if pd.notna(row_data['Chi-Square']) else '', highlight=is_sig)
            write_data_cell(ws, cur, 4, round(float(row_data['Pr(>Chisq)']), 4) if pd.notna(row_data['Pr(>Chisq)']) else '', highlight=is_sig)
            write_data_cell(ws, cur, 5, row_data.get('Signif', ''), highlight=is_sig)
            cur += 1
    cur += 2

    ws.cell(row=cur, column=1, value="【2. 回帰係数とオッズ比】").font = subhead_font
    cur += 1
    headers_g = ["要因 / 水準", "Estimate", "Odds Ratio", "Std.Error", "z value", "Pr(>|z|)", "有意"]
    write_header_row(ws, cur, headers_g)
    cur += 1
    for str_idx, row_data in report_glm.iterrows():
        is_sig = float(row_data['Pr(>|z|)']) < ALPHA if pd.notna(row_data['Pr(>|z|)']) else False
        write_data_cell(ws, cur, 1, str(str_idx), highlight=is_sig)
        write_data_cell(ws, cur, 2, round(float(row_data['Estimate']), 4) if pd.notna(row_data['Estimate']) else '', highlight=is_sig)
        write_data_cell(ws, cur, 3, round(float(row_data['Odds Ratio']), 4) if pd.notna(row_data['Odds Ratio']) else '', highlight=is_sig)
        write_data_cell(ws, cur, 4, round(float(row_data['Std.Error']), 4) if pd.notna(row_data['Std.Error']) else '', highlight=is_sig)
        write_data_cell(ws, cur, 5, round(float(row_data['z value']), 3) if pd.notna(row_data['z value']) else '', highlight=is_sig)
        write_data_cell(ws, cur, 6, round(float(row_data['Pr(>|z|)']), 4) if pd.notna(row_data['Pr(>|z|)']) else '', highlight=is_sig)
        write_data_cell(ws, cur, 7, row_data.get('Signif', '') if not pd.isna(row_data['Std.Error']) and row_data['Std.Error'] <= 10 else '⚠️SE過大', highlight=is_sig)
        cur += 1
    cur += 3

    if report_images:
        ws.cell(row=cur, column=1, value="【3. 項目別の影響グラフ】").font = subhead_font
        cur += 2
        for factor_name, img_buf in report_images.items():
            ws.cell(row=cur, column=1, value=f"▶ {factor_name}").font = Font(bold=True)
            cur += 1
            img_buf.seek(0)
            try:
                xl_img = OpenpyxlImage(img_buf)
                xl_img.width  = int(xl_img.width  * 0.75)
                xl_img.height = int(xl_img.height * 0.75)
                ws.add_image(xl_img, f"B{cur}")
                rows_to_skip = max(int(xl_img.height / 18) + 2, 5)
            except Exception:
                ws.cell(row=cur, column=1, value="（グラフ挿入失敗）")
                rows_to_skip = 2
            cur += rows_to_skip

    ws.column_dimensions['A'].width = 35
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']: ws.column_dimensions[col_letter].width = 14
    wb.save(output)
    return output.getvalue()


# ==========================================
# アプリ本体
# ==========================================
st.title("📋 アンケートデータ GLM解析アプリ")
st.markdown("質問項目（質的/量的データ）から「はい/いいえ」「購入する/しない」といった2値の要因を解析する、**ロジスティック回帰（二項分布）**特化型アプリです。")

if FONT_STATUS == "fallback":
    st.warning("⚠️ 日本語フォントが見つかりませんでした。文字化けする場合は `pip install japanize-matplotlib` を実行してください。")

with st.sidebar:
    st.header("⚙️ 操作")
    if st.button("🔄 リセット（最初からやり直す）", use_container_width=True):
        reset_session()
    st.divider()
    st.caption(f"有意水準: α = {ALPHA}")

# ==========================================
# Step 1: データ読み込み
# ==========================================
st.header("1. データの読み込み")
data_source = st.radio(
    "入力方法を選択：",
    ["📄 ファイルアップロード", "📋 Excelデータを貼り付け", "📊 官能評価サンプルで試す"],
    horizontal=True
)
df_raw = None

if data_source == "📄 ファイルアップロード":
    uploaded_file = st.file_uploader("CSV または Excelファイルをアップロードしてください", type=["csv", "xlsx"])
    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith('.csv'):
                try:
                    df_raw = pd.read_csv(uploaded_file)
                except UnicodeDecodeError:
                    uploaded_file.seek(0)
                    df_raw = pd.read_csv(uploaded_file, encoding='shift_jis')
            else:
                df_raw = pd.read_excel(uploaded_file)
            st.success(f"✅ ファイルを読み込みました（{len(df_raw)} 件）。")
        except Exception as e:
            st.error(f"❌ 読み込みエラー: {e}")

elif data_source == "📊 官能評価サンプルで試す":
    # アンケート向けのダミーデータ生成
    rng = np.random.default_rng(42)
    rows = []
    for _ in range(250):
        var = rng.choice(['シマアカリ', 'ニシユタカ', 'アイユタカ'])
        age = rng.choice(['20代', '30代', '40代', '50代以上'], p=[0.2, 0.3, 0.3, 0.2])
        texture = rng.integers(1, 6) # 1〜5の評価
        sweetness = rng.integers(1, 6)
        
        # ロジスティック回帰の背後にある確率を生成
        base_logit = -3.0 + 0.8 * texture + 0.5 * sweetness
        if var == 'シマアカリ': base_logit += 1.2
        if age == '20代': base_logit -= 0.8
        
        prob = 1 / (1 + np.exp(-base_logit))
        purchase = rng.binomial(1, prob)
        rows.append([var, age, texture, sweetness, purchase])
        
    df_raw = pd.DataFrame(rows, columns=['試食品種', '回答者年代', '食感評価(1-5)', '甘み評価(1-5)', '購入意向(1=買う)'])
    st.success("✅ サンプルデータを読み込みました。")

else:
    col_paste1, col_paste2 = st.columns([3, 1])
    with col_paste2:
        sep_choice = st.selectbox("区切り文字", ["自動検出", "タブ (\\t)", "カンマ (,)", "スペース"])
    with col_paste1:
        pasted_data = st.text_area("Excel/CSVデータを貼り付け (Ctrl+V)", height=100)

    if pasted_data:
        sep_map = {"自動検出": None, "タブ (\\t)": '\t', "カンマ (,)": ',', "スペース": r'\s+'}
        sep_val = sep_map[sep_choice]
        try:
            if sep_val is None:
                df_raw = pd.read_csv(io.StringIO(pasted_data), sep=None, engine='python')
            elif sep_val == r'\s+':
                df_raw = pd.read_csv(io.StringIO(pasted_data), sep=r'\s+', engine='python')
            else:
                df_raw = pd.read_csv(io.StringIO(pasted_data), sep=sep_val)
            st.success(f"✅ データを読み込みました（{len(df_raw)} 件）。")
        except Exception as e:
            st.error(f"読み込みエラー: {e}")

if df_raw is None:
    st.stop()

# ==========================================
# Step 2 & 3: データの確認
# ==========================================
st.header("2. データの確認")
col_s1, col_s2 = st.columns(2)
with col_s1:
    st.markdown("**データプレビュー**")
    st.dataframe(df_raw.head(5), use_container_width=True)
with col_s2:
    st.markdown("**要約統計量**")
    st.dataframe(df_raw.describe(), use_container_width=True)

st.divider()

# ==========================================
# Step 4: モデル設定
# ==========================================
st.header("3. モデル設定")
cols = df_raw.columns.tolist()
numeric_cols = [c for c in cols if pd.api.types.is_numeric_dtype(df_raw[c])]

col_m1, col_m2 = st.columns(2)
with col_m1:
    st.markdown("**目的変数（回答結果）**")
    target_col = st.selectbox("「はい/いいえ」「1/0」を示す列を選択", numeric_cols)
    st.info("💡 分析手法は自動的に**二項分布（ロジスティック回帰）**に設定されます。")

with col_m2:
    st.markdown("**要因（質問項目・回答者属性）**")
    all_factor_candidates = [c for c in cols if c != target_col]
    factor_cols = st.multiselect("分析に含める要因列を選択", all_factor_candidates, default=all_factor_candidates[:3])

factor_types = {}
if factor_cols:
    st.markdown("**各要因のデータ型：**")
    for f in factor_cols:
        is_text = df_raw[f].dtype == object
        default_type = "カテゴリ (属性や選択肢)" if is_text or df_raw[f].nunique() <= 10 else "数値 (スコアや年齢)"
        factor_types[f] = st.radio(f"【{f}】", ["カテゴリ (属性や選択肢)", "数値 (スコアや年齢)"], index=0 if "カテゴリ" in default_type else 1, horizontal=True, key=f"type_{f}")

run_button = st.button("🚀 ロジスティック回帰を実行", type="primary", disabled=(len(factor_cols) == 0))

# ==========================================
# Step 5: GLM実行
# ==========================================
if run_button and factor_cols and target_col:
    # 古いキャッシュをクリア
    for k in ['report_images', 'model_result', 'report_glm', 'report_wald', 'df_eval', 'eval_col', 'formula', 'excel_data']:
        if k in st.session_state: del st.session_state[k]

    st.divider()
    st.header("4. 解析結果")

    # データの前処理と欠損値ハンドリング
    df_eval = df_raw.copy()
    n_before = len(df_eval)
    df_eval[target_col] = safe_numeric(df_eval[target_col])
    
    # 欠損値（無回答）を含む行を削除（リストワイズ除去）
    df_eval = df_eval.dropna(subset=[target_col] + factor_cols)
    n_after = len(df_eval)
    if n_after < n_before:
        st.warning(f"⚠️ 欠損値（無回答）を含むデータを **{n_before - n_after} 件** 除外しました。解析対象: {n_after} 件。")

    if df_eval.empty: st.error("⚠️ データが0件になりました。"); st.stop()

    eval_col = f"{target_col}_bin"
    # 二値データ(0,1)へのクリッピング
    df_eval[eval_col] = np.clip(df_eval[target_col], 0, 1)

    # 完全分離の事前チェック
    sep_warnings = []
    for f in factor_cols:
        if "カテゴリ" in factor_types[f]:
            prob_cats = detect_perfect_separation(df_eval, eval_col, f)
            if prob_cats:
                sep_warnings.append(f"**{f}**: 選択肢「{', '.join(prob_cats)}」で回答が全員一致しています（完全分離）。結果の信頼性が落ちる可能性があります。")
    if sep_warnings:
        with st.expander("⚠️ データ偏りの警告（完全分離）", expanded=True):
            for w in sep_warnings: st.warning(w)

    # モデル式の構築
    formula_terms = []
    for f in factor_cols:
        if "カテゴリ" in factor_types[f]:
            df_eval[f] = df_eval[f].astype(str)
            formula_terms.append(f'C(Q("{f}"))')
        else:
            df_eval[f] = safe_numeric(df_eval[f])
            formula_terms.append(f'Q("{f}")')

    formula = f'Q("{eval_col}") ~ ' + ' + '.join(formula_terms)

    try:
        with st.spinner("モデルを計算中..."):
            # ロジスティック回帰の実行
            family = sm.families.Binomial(link=sm.families.links.logit())
            model = smf.glm(formula, data=df_eval, family=family).fit()

        if model.df_resid <= 0:
            st.error("❌ データ数に対して要因の数が多すぎます。")
            st.stop()

        # サマリーテーブルの作成
        report_glm = pd.DataFrame({
            'Estimate':  model.params,
            'Std.Error': model.bse,
            'z value':   model.tvalues,
            'Pr(>|z|)':  model.pvalues
        })
        # ★オッズ比の計算を追加
        report_glm.insert(1, 'Odds Ratio', np.exp(report_glm['Estimate']))
        
        report_glm['Signif'] = report_glm['Pr(>|z|)'].apply(lambda p: "**" if p < 0.01 else ("*" if p < ALPHA else "n.s."))
        report_glm['Note'] = report_glm['Std.Error'].apply(lambda se: "⚠️SE過大" if pd.notna(se) and se > 10 else "")

        # Wald検定表の作成
        report_wald = None
        try:
            wald_res = model.wald_test_terms().table
            report_wald = wald_res[['df_constraint', 'statistic', 'pvalue']].copy()
            report_wald.columns = ['Df', 'Chi-Square', 'Pr(>Chisq)']
            for col_w in report_wald.columns:
                report_wald[col_w] = pd.to_numeric(report_wald[col_w], errors='coerce').astype(float)
            report_wald['Signif'] = report_wald['Pr(>Chisq)'].apply(lambda p: "**" if p < 0.01 else ("*" if p < ALPHA else "n.s."))
        except Exception:
            pass

        st.session_state['report_glm']  = report_glm
        st.session_state['report_wald'] = report_wald
        st.session_state['df_eval']     = df_eval
        st.session_state['eval_col']    = eval_col
        st.session_state['formula']     = formula
        st.session_state['target_col']  = target_col
        st.session_state['factor_cols'] = factor_cols
        st.session_state['factor_types']= factor_types
        st.session_state['analyzed']    = True

    except PerfectSeparationError:
        st.error("❌ **完全分離エラー**\n\n特定の回答者グループで結果が完全に偏っています。要因を減らしてください。")
        st.stop()
    except Exception as e:
        st.error(f"❌ モデル計算エラー: {e}")
        st.stop()


# ==========================================
# Step 6以降: セッション状態から結果を表示
# ==========================================
if st.session_state.get('analyzed'):

    report_glm   = st.session_state['report_glm']
    report_wald  = st.session_state['report_wald']
    df_eval      = st.session_state['df_eval']
    eval_col     = st.session_state['eval_col']
    formula      = st.session_state['formula']
    target_col   = st.session_state['target_col']
    factor_cols  = st.session_state['factor_cols']
    factor_types = st.session_state['factor_types']

    tab_res, tab_ai = st.tabs(["📊 解析サマリー", "🤖 AI解析用プロンプト"])

    with tab_res:
        st.markdown("""
        **💡 読み方のポイント**
        * **オッズ比 (Odds Ratio)**: `1.0` を基準とします。`1.5`なら「該当しやすさが1.5倍に上がる」、`0.5`なら「該当しやすさが半分に下がる」と解釈できます。
        """)
        
        col_r1, col_r2 = st.columns([3, 2])
        with col_r1:
            st.markdown("**■ 回帰係数とオッズ比**")
            styled_df = report_glm.style.format({
                'Estimate': '{:.4f}', 'Odds Ratio': '{:.3f}', 'Std.Error': '{:.4f}',
                'z value': '{:.3f}', 'Pr(>|z|)': '{:.4f}'
            })
            try:
                styled_df = styled_df.map(lambda v: 'background-color: #fff3cd' if '⚠️' in str(v) else '', subset=['Note'])
            except AttributeError:
                styled_df = styled_df.applymap(lambda v: 'background-color: #fff3cd' if '⚠️' in str(v) else '', subset=['Note'])
            st.dataframe(styled_df, use_container_width=True)

        with col_r2:
            st.markdown("**■ Wald検定表 (質問項目全体の有意差)**")
            if report_wald is not None:
                st.dataframe(
                    report_wald.style.format({
                        'Df': '{:.0f}', 'Chi-Square': '{:.2f}', 'Pr(>Chisq)': '{:.4f}'
                    }),
                    use_container_width=True
                )
            else:
                st.info("計算失敗")

    with tab_ai:
        st.markdown("### AIへの解析指示プロンプト")
        st.caption("ChatGPTなどに貼り付けて、解釈を手伝ってもらえます。")
        ai_prompt = generate_ai_prompt(target_col, formula, report_glm, report_wald)
        st.text_area("📋 プロンプトをコピー", value=ai_prompt, height=300)

    st.divider()

    st.header("5. 質問項目別の影響グラフ")
    if 'report_images' not in st.session_state:
        st.session_state['report_images'] = {}

    for factor in factor_cols:
        is_cat = "カテゴリ" in factor_types[factor]
        st.subheader(f"▶ {factor} {'(選択肢)' if is_cat else '(スコア・数値)'}")

        col_n1, col_n2 = st.columns([2, 3])
        with col_n1:
            if is_cat:
                summary_stats = df_eval.groupby(factor)[eval_col].agg(['count', 'mean']).reset_index()
                summary_stats.columns = [factor, '回答数(N)', '該当割合(%)']
                summary_stats['該当割合(%)'] = (summary_stats['該当割合(%)'] * 100).round(1)
                st.dataframe(summary_stats, use_container_width=True)
            else:
                st.dataframe(df_eval[factor].describe(), use_container_width=True)
                
        with col_n2:
            if is_cat:
                fig = make_fig_for_category(df_eval, factor, eval_col, target_col)
            else:
                fig = make_fig_for_numeric(df_eval, factor, eval_col, target_col)
            st.pyplot(fig)
            st.session_state['report_images'][factor] = fig_to_bytesio(fig)

    st.divider()

    st.header("6. 解析レポートのダウンロード")
    if 'excel_data' not in st.session_state:
        with st.spinner("Excelレポートを生成中..."):
            try:
                st.session_state['excel_data'] = generate_excel_report(
                    target_col, formula, report_glm, report_wald, st.session_state.get('report_images', {})
                )
            except Exception as e:
                st.error(f"Excelレポート生成エラー: {e}")
                st.session_state['excel_data'] = None

    if st.session_state.get('excel_data'):
        st.download_button(
            label="📥 解析レポート（図入りExcel）をダウンロード",
            data=st.session_state['excel_data'],
            file_name="Questionnaire_GLM_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
